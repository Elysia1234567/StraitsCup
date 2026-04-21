#!/usr/bin/env python3
"""
将 Util/standardList.xlsx 转换为适合 RAG 入库的标准化数据。

输出：
- JSONL：每行一个标准化知识条目，便于后续切块、向量化、入库

默认读取：
- Util/standardList.xlsx

默认输出：
- Util/standardList.jsonl

示例：
    python Util/import_standard_list.py
    python Util/import_standard_list.py --input Util/standardList.xlsx --output Util/standardList.jsonl --dedupe
"""

from __future__ import annotations

import argparse
import json
import logging
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook


LOGGER = logging.getLogger(__name__)


FIELD_ALIASES = {
    "项目名称": ["项目名称", "名称", "题名", "项目名"],
    "非遗级别": ["非遗级别", "级别", "等级"],
    "类别": ["类别", "项目类别", "分类"],
    "所属地区": ["所属地区", "地区", "所在地", "申报地区"],
    "简介": ["简介", "内容", "描述", "项目简介"],
    "代表作品": ["代表作品", "作品", "代表作", "示例作品"],
}


@dataclass
class StandardRecord:
    """标准化后的知识条目。"""

    id: str
    title: str
    content: str
    metadata: dict[str, Any]


def normalize_text(value: Any) -> str:
    """清洗单元格文本。"""
    if value is None:
        return ""
    text = str(value)
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    lines = [line.strip() for line in text.split("\n") if line.strip()]
    return "；".join(lines).strip()


def build_header_map(headers: list[Any]) -> dict[str, int]:
    """把表头映射为列索引。"""
    header_map: dict[str, int] = {}
    for index, header in enumerate(headers):
        if header is None:
            continue
        header_map[str(header).strip()] = index
    return header_map


def find_column(header_map: dict[str, int], aliases: Iterable[str]) -> int | None:
    """按别名寻找列索引。"""
    for alias in aliases:
        if alias in header_map:
            return header_map[alias]
    return None


def make_record_id(sheet_name: str, row_index: int, title: str) -> str:
    """生成稳定 ID。"""
    base = title.strip() if title.strip() else f"row-{row_index}"
    return f"{sheet_name}-{row_index}-{base}"


def build_content(fields: dict[str, str]) -> str:
    """构造适合 RAG 检索的正文。"""
    parts: list[str] = []
    if fields.get("title"):
        parts.append(f"项目名称：{fields['title']}")
    if fields.get("level"):
        parts.append(f"非遗级别：{fields['level']}")
    if fields.get("category"):
        parts.append(f"类别：{fields['category']}")
    if fields.get("region"):
        parts.append(f"所属地区：{fields['region']}")
    if fields.get("intro"):
        parts.append(f"简介：{fields['intro']}")
    if fields.get("works"):
        parts.append(f"代表作品：{fields['works']}")
    return "\n".join(parts).strip()


def parse_workbook(input_path: Path, sheet_name: str | None = None) -> list[StandardRecord]:
    """读取 Excel 并转换为标准化记录。"""
    wb = load_workbook(input_path, data_only=True)
    records: list[StandardRecord] = []

    target_sheets = [sheet_name] if sheet_name else wb.sheetnames
    for current_sheet_name in target_sheets:
        if current_sheet_name not in wb.sheetnames:
            raise ValueError(f"工作表不存在: {current_sheet_name}")

        ws = wb[current_sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            LOGGER.warning("工作表 %s 为空，跳过。", current_sheet_name)
            continue

        header_map = build_header_map(list(rows[0]))
        title_col = find_column(header_map, FIELD_ALIASES["项目名称"])
        level_col = find_column(header_map, FIELD_ALIASES["非遗级别"])
        category_col = find_column(header_map, FIELD_ALIASES["类别"])
        region_col = find_column(header_map, FIELD_ALIASES["所属地区"])
        intro_col = find_column(header_map, FIELD_ALIASES["简介"])
        works_col = find_column(header_map, FIELD_ALIASES["代表作品"])

        for row_idx, row in enumerate(rows[1:], start=2):
            title = normalize_text(row[title_col]) if title_col is not None and title_col < len(row) else ""
            level = normalize_text(row[level_col]) if level_col is not None and level_col < len(row) else ""
            category = normalize_text(row[category_col]) if category_col is not None and category_col < len(row) else ""
            region = normalize_text(row[region_col]) if region_col is not None and region_col < len(row) else ""
            intro = normalize_text(row[intro_col]) if intro_col is not None and intro_col < len(row) else ""
            works = normalize_text(row[works_col]) if works_col is not None and works_col < len(row) else ""

            if not any([title, level, category, region, intro, works]):
                continue

            fields = {
                "title": title,
                "level": level,
                "category": category,
                "region": region,
                "intro": intro,
                "works": works,
            }
            record = StandardRecord(
                id=make_record_id(current_sheet_name, row_idx, title),
                title=title or f"未命名-{row_idx}",
                content=build_content(fields),
                metadata={
                    "sheet": current_sheet_name,
                    "row": row_idx,
                    "source_file": input_path.name,
                    **fields,
                },
            )
            records.append(record)

    return records


def dedupe_records(records: list[StandardRecord]) -> list[StandardRecord]:
    """按 title + level + region 去重。"""
    seen: set[tuple[str, str, str]] = set()
    deduped: list[StandardRecord] = []
    for record in records:
        meta = record.metadata
        key = (
            (meta.get("title") or record.title).strip(),
            (meta.get("level") or "").strip(),
            (meta.get("region") or "").strip(),
        )
        if key in seen:
            continue
        seen.add(key)
        deduped.append(record)
    return deduped


def write_jsonl(records: list[StandardRecord], output_path: Path) -> None:
    """写出 JSONL 文件。"""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8") as f:
        for record in records:
            f.write(json.dumps(asdict(record), ensure_ascii=False) + "\n")


def main() -> int:
    parser = argparse.ArgumentParser(description="将 standardList.xlsx 转换为 RAG 可入库的 JSONL。")
    parser.add_argument(
        "--input",
        default=str(Path(__file__).with_name("standardList.xlsx")),
        help="输入 Excel 路径，默认 Util/standardList.xlsx",
    )
    parser.add_argument(
        "--output",
        default=str(Path(__file__).with_name("standardList.jsonl")),
        help="输出 JSONL 路径，默认 Util/standardList.jsonl",
    )
    parser.add_argument(
        "--sheet",
        default=None,
        help="仅处理指定工作表；不填则处理所有工作表",
    )
    parser.add_argument(
        "--dedupe",
        action="store_true",
        help="按 title + level + region 去重",
    )
    parser.add_argument(
        "--preview",
        type=int,
        default=3,
        help="打印前几条记录预览",
    )
    parser.add_argument("--verbose", action="store_true")

    args = parser.parse_args()
    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(levelname)s: %(message)s",
    )

    input_path = Path(args.input).expanduser().resolve()
    output_path = Path(args.output).expanduser().resolve()

    if not input_path.exists():
        LOGGER.error("输入文件不存在: %s", input_path)
        return 1

    LOGGER.info("开始读取: %s", input_path)
    records = parse_workbook(input_path, sheet_name=args.sheet)

    if args.dedupe:
        before = len(records)
        records = dedupe_records(records)
        LOGGER.info("去重完成：%d -> %d", before, len(records))

    if not records:
        LOGGER.warning("没有解析到任何有效记录。")
        return 0

    write_jsonl(records, output_path)
    LOGGER.info("已写出 %d 条记录到 %s", len(records), output_path)

    preview_count = min(max(args.preview, 0), len(records))
    if preview_count:
        LOGGER.info("预览前 %d 条记录：", preview_count)
        for record in records[:preview_count]:
            LOGGER.info("- %s", json.dumps(asdict(record), ensure_ascii=False))

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
